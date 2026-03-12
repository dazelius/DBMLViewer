"""
바이블테이블링 — Excel 편집 엔진 (openpyxl 기반)

핵심 기능:
- 게임 데이터 xlsx 파일 읽기/편집
- AI가 편집한 셀은 노란색(#FFFF00) 하이라이트
- 필터 조건으로 대상 행 선택
- set/multiply/add/subtract/append 변경 액션
"""

import csv
import io
import shutil
import time
import xml.etree.ElementTree as ET
import zipfile
from collections import OrderedDict
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ── AI 편집 표시용 노란색 하이라이트 ──
AI_HIGHLIGHT = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

# ── 메타 시트 (편집 대상에서 제외) ──
META_SHEET_NAMES = {'define', 'enum', 'tablegroup', 'ref', 'tabledefine', 'sheet1'}


def _fast_sheet_names(xlsx_path: Path) -> list[str]:
    """zipfile로 시트 이름만 빠르게 추출 (openpyxl load_workbook보다 10x+ 빠름)"""
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            with z.open('xl/workbook.xml') as f:
                tree = ET.parse(f)
                ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                return [sheet.get('name', '') for sheet in tree.findall('.//s:sheet', ns)]
    except Exception:
        try:
            wb = load_workbook(xlsx_path, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception:
            return []


class BibleTablingEditor:
    """게임 데이터 Excel 편집기"""

    def __init__(self, data_dir: str):
        self.data_dir = Path(data_dir)
        if not self.data_dir.exists():
            raise ValueError(f"데이터 디렉토리가 없습니다: {data_dir}")
        self._sheet_cache: dict[str, tuple[float, list[str]]] = {}  # path → (mtime, sheet_names)

    # ── 테이블 목록 ──────────────────────────────────────────────────────────

    def list_tables(self) -> list[dict]:
        """사용 가능한 테이블 목록 반환"""
        tables = []
        for xlsx_path in sorted(self.data_dir.glob("*.xlsx")):
            if xlsx_path.name.startswith("~$"):
                continue
            try:
                wb = load_workbook(xlsx_path, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    if sheet_name.lower() in META_SHEET_NAMES:
                        continue
                    if '#' in sheet_name:
                        continue
                    ws = wb[sheet_name]
                    # 헤더 행 찾기
                    headers = []
                    for row in ws.iter_rows(max_row=3, values_only=True):
                        non_empty = [str(c) for c in row if c is not None]
                        if len(non_empty) > len(headers):
                            headers = non_empty
                    row_count = ws.max_row - 1 if ws.max_row else 0
                    tables.append({
                        "table": sheet_name,
                        "file": xlsx_path.name,
                        "columns": headers[:20],
                        "row_count": max(0, row_count),
                    })
                wb.close()
            except Exception:
                continue
        return tables

    # ── 파일 탐색 ────────────────────────────────────────────────────────────

    def _get_sheet_names(self, xlsx_path: Path) -> list[str]:
        """시트 이름 캐시 (mtime 기반 무효화)"""
        key = str(xlsx_path)
        try:
            mtime = xlsx_path.stat().st_mtime
        except OSError:
            return []
        cached = self._sheet_cache.get(key)
        if cached and cached[0] == mtime:
            return cached[1]
        names = _fast_sheet_names(xlsx_path)
        self._sheet_cache[key] = (mtime, names)
        return names

    def _find_xlsx_file(self, table: str, file_hint: str = None) -> Path:
        """테이블 이름으로 xlsx 파일 찾기 (캐시 활용)"""
        if file_hint:
            path = self.data_dir / file_hint
            if path.exists():
                return path

        for xlsx_path in self.data_dir.glob("*.xlsx"):
            if xlsx_path.name.startswith("~$"):
                continue
            if xlsx_path.stem.lower() == table.lower():
                return xlsx_path

        for xlsx_path in self.data_dir.glob("*.xlsx"):
            if xlsx_path.name.startswith("~$"):
                continue
            sheet_names = [sn.lower() for sn in self._get_sheet_names(xlsx_path)]
            if table.lower() in sheet_names:
                return xlsx_path

        raise FileNotFoundError(f"테이블 '{table}'에 해당하는 xlsx 파일을 찾을 수 없습니다.")

    # ── 헤더 파싱 ────────────────────────────────────────────────────────────

    def _find_header_row(self, ws) -> int:
        """헤더 행 인덱스 찾기 (1-based)"""
        best_idx = 1
        best_count = 0
        for row_idx in range(1, min(5, ws.max_row + 1)):
            count = sum(
                1 for cell in ws[row_idx]
                if cell.value is not None and str(cell.value).strip()
            )
            if count > best_count:
                best_count = count
                best_idx = row_idx
        return best_idx

    def _get_header_map(self, ws, header_row: int) -> dict[str, int]:
        """헤더명 → 컬럼 인덱스 매핑 (1-based)"""
        header_map = {}
        for cell in ws[header_row]:
            if cell.value is not None:
                name = str(cell.value).strip()
                if name:
                    header_map[name.lower()] = cell.column
                    header_map[name] = cell.column  # 원본 케이스도 등록
        return header_map

    # ── 필터 매칭 ────────────────────────────────────────────────────────────

    @staticmethod
    def _matches_filter(cell_value, filter_cond) -> bool:
        """셀 값이 필터 조건에 매치하는지 확인"""
        val = str(cell_value).strip() if cell_value is not None else ''
        op = filter_cond.op.lower()

        if op == 'eq':
            return val.lower() == str(filter_cond.value).lower()
        elif op == 'neq':
            return val.lower() != str(filter_cond.value).lower()
        elif op == 'in':
            target_values = filter_cond.values or (
                [str(filter_cond.value)] if filter_cond.value else []
            )
            return val.lower() in [str(v).lower() for v in target_values]
        elif op == 'contains':
            return str(filter_cond.value).lower() in val.lower()
        elif op == 'starts_with':
            return val.lower().startswith(str(filter_cond.value).lower())
        elif op == 'ends_with':
            return val.lower().endswith(str(filter_cond.value).lower())
        elif op in ('gt', 'gte', 'lt', 'lte'):
            try:
                num_val = float(val)
                num_target = float(filter_cond.value)
                if op == 'gt': return num_val > num_target
                if op == 'gte': return num_val >= num_target
                if op == 'lt': return num_val < num_target
                if op == 'lte': return num_val <= num_target
            except (ValueError, TypeError):
                return False

        return False

    # ── 셀 변경 적용 ────────────────────────────────────────────────────────

    @staticmethod
    def _apply_change(cell, change) -> bool:
        """셀에 변경 적용 + 노란색 하이라이트. 성공 시 True."""
        action = change.action.lower()
        old_value = cell.value

        if action == 'set':
            # 숫자처럼 보이면 숫자로 저장
            try:
                v = float(change.value)
                cell.value = int(v) if v == int(v) else round(v, 4)
            except (ValueError, TypeError):
                cell.value = change.value
        elif action in ('multiply', 'add', 'subtract'):
            try:
                current = float(old_value) if old_value is not None else 0
                factor = float(change.value)
                if action == 'multiply':
                    new_val = current * factor
                elif action == 'add':
                    new_val = current + factor
                else:  # subtract
                    new_val = current - factor

                # 정수면 정수, 아니면 소수점 4자리
                cell.value = int(new_val) if new_val == int(new_val) else round(new_val, 4)
            except (ValueError, TypeError):
                return False
        elif action == 'append':
            current = str(old_value) if old_value is not None else ''
            cell.value = current + str(change.value)
        else:
            return False

        # ── 노란색 하이라이트 적용 ──
        cell.fill = AI_HIGHLIGHT
        return True

    # ── 단일 테이블 편집 (워크북이 이미 열려 있으면 재사용) ─────────────────────

    # ── CSV 기반 대량 set 편집 (PK 룩업 O(n+m)) ────────────────────────────

    def apply_csv_bulk_set(self, edit, wb, xlsx_path: Path) -> dict:
        """
        CSV 기반 대량 set 편집. 첫 번째 열 = PK 필터(eq), 나머지 열 = set 값.
        PK→row 룩업 테이블을 한 번 구축하여 O(n+m) 성능.
        """
        sheet_name = edit.sheet or edit.table
        ws = None
        for sn in wb.sheetnames:
            if sn.lower() == sheet_name.lower():
                ws = wb[sn]
                break
        if ws is None:
            raise ValueError(
                f"시트 '{sheet_name}' 없음 ({xlsx_path.name}, 시트: {wb.sheetnames})"
            )

        header_row = self._find_header_row(ws)
        header_map = self._get_header_map(ws, header_row)

        reader = csv.reader(io.StringIO(edit.csv_set.strip()))
        csv_headers = [h.strip() for h in next(reader)]
        if len(csv_headers) < 2:
            raise ValueError("csv_set에 최소 2열(PK + 변경열) 필요")

        pk_col_name = csv_headers[0]
        set_col_names = csv_headers[1:]

        pk_col_idx = header_map.get(pk_col_name) or header_map.get(pk_col_name.lower())
        if pk_col_idx is None:
            raise ValueError(f"PK 컬럼 '{pk_col_name}' 없음 (시트: {sheet_name})")

        set_col_indices = []
        for cn in set_col_names:
            idx = header_map.get(cn) or header_map.get(cn.lower())
            if idx is None:
                raise ValueError(f"컬럼 '{cn}' 없음 (시트: {sheet_name})")
            set_col_indices.append(idx)

        pk_row_map: dict[str, int] = {}
        for row_idx in range(header_row + 1, ws.max_row + 1):
            pk_val = ws.cell(row=row_idx, column=pk_col_idx).value
            if pk_val is not None:
                pk_row_map[str(pk_val).strip().lower()] = row_idx

        cells_modified = 0
        rows_matched = 0
        change_details = []

        for csv_row in reader:
            if not csv_row or not csv_row[0].strip():
                continue
            pk_value = csv_row[0].strip()
            row_idx = pk_row_map.get(pk_value.lower())
            if row_idx is None:
                continue

            rows_matched += 1
            for i, col_idx in enumerate(set_col_indices):
                if i + 1 >= len(csv_row):
                    continue
                new_val_str = csv_row[i + 1]
                cell = ws.cell(row=row_idx, column=col_idx)
                old_val = cell.value

                try:
                    v = float(new_val_str)
                    cell.value = int(v) if v == int(v) else round(v, 4)
                except (ValueError, TypeError):
                    cell.value = new_val_str

                cell.fill = AI_HIGHLIGHT
                cells_modified += 1
                change_details.append({
                    "row": row_idx,
                    "pk": pk_value,
                    "column": set_col_names[i],
                    "old": str(old_val) if old_val is not None else "",
                    "new": str(cell.value),
                })

        return {
            "table": edit.table,
            "file": xlsx_path.name,
            "sheet": sheet_name,
            "rows_matched": rows_matched,
            "cells_modified": cells_modified,
            "changes": change_details[:100],
        }

    # ── 단일 테이블 편집 (워크북이 이미 열려 있으면 재사용) ─────────────────────

    def apply_edit_to_wb(self, edit, wb, xlsx_path: Path) -> dict:
        """
        이미 열린 워크북에 대해 단일 시트 편집을 적용.
        wb.save()는 호출자가 모든 편집 완료 후 한 번만 수행.
        """
        sheet_name = edit.sheet or edit.table

        ws = None
        for sn in wb.sheetnames:
            if sn.lower() == sheet_name.lower():
                ws = wb[sn]
                break

        if ws is None:
            raise ValueError(
                f"시트 '{sheet_name}'을 찾을 수 없습니다. "
                f"(파일: {xlsx_path.name}, 시트 목록: {wb.sheetnames})"
            )

        header_row = self._find_header_row(ws)
        header_map = self._get_header_map(ws, header_row)

        matched_rows = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            first_cell = ws.cell(row=row_idx, column=1).value
            if first_cell is None:
                continue

            all_match = True
            for f in edit.filters:
                col_idx = header_map.get(f.column) or header_map.get(f.column.lower())
                if col_idx is None:
                    all_match = False
                    break
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if not self._matches_filter(cell_value, f):
                    all_match = False
                    break

            if all_match:
                matched_rows.append(row_idx)

        cells_modified = 0
        change_details = []

        for row_idx in matched_rows:
            for change in edit.changes:
                col_idx = header_map.get(change.column) or header_map.get(change.column.lower())
                if col_idx is None:
                    continue

                cell = ws.cell(row=row_idx, column=col_idx)
                old_val = cell.value

                if self._apply_change(cell, change):
                    cells_modified += 1
                    pk_val = ws.cell(row=row_idx, column=1).value
                    change_details.append({
                        "row": row_idx,
                        "pk": str(pk_val) if pk_val is not None else "",
                        "column": change.column,
                        "old": str(old_val) if old_val is not None else "",
                        "new": str(cell.value),
                    })

        return {
            "table": edit.table,
            "file": xlsx_path.name,
            "sheet": sheet_name,
            "rows_matched": len(matched_rows),
            "cells_modified": cells_modified,
            "changes": change_details[:100],
        }

    def apply_edit(self, edit, output_dir: Path, prev_job_dir: Path = None) -> dict:
        """단일 테이블 편집 (하위 호환용 래퍼)"""
        xlsx_path = self._find_xlsx_file(edit.table, edit.file)
        output_path = output_dir / xlsx_path.name
        if not output_path.exists():
            prev_path = prev_job_dir / xlsx_path.name if prev_job_dir else None
            if prev_path and prev_path.exists():
                shutil.copy2(prev_path, output_path)
            else:
                shutil.copy2(xlsx_path, output_path)

        wb = load_workbook(output_path)
        result = self.apply_edit_to_wb(edit, wb, xlsx_path)
        wb.save(output_path)
        wb.close()
        return result

    def _group_edits_by_file(self, edits: list) -> OrderedDict[str, dict]:
        file_groups: OrderedDict[str, dict] = OrderedDict()
        for edit in edits:
            xlsx_path = self._find_xlsx_file(edit.table, edit.file)
            fname = xlsx_path.name
            if fname not in file_groups:
                file_groups[fname] = {"xlsx_path": xlsx_path, "edits": []}
            file_groups[fname]["edits"].append(edit)
        return file_groups

    def _execute_single_edit(self, edit, wb, xlsx_path: Path) -> dict:
        """단일 편집 실행 (csv_set/기존 방식 자동 분기)"""
        if hasattr(edit, 'csv_set') and edit.csv_set:
            return self.apply_csv_bulk_set(edit, wb, xlsx_path)
        return self.apply_edit_to_wb(edit, wb, xlsx_path)

    def apply_edits_batch(self, edits: list, output_dir: Path, prev_job_dir: Path = None) -> list[dict]:
        """
        여러 편집을 파일별로 그룹화하여 한 번만 열고/저장.
        같은 xlsx의 여러 시트를 편집할 때 이전 시트 변경이 유실되지 않음.
        개별 편집 실패 시에도 나머지를 계속 처리 (fault-tolerant).
        """
        file_groups = self._group_edits_by_file(edits)

        results = []
        t_total = time.time()
        for fname, group in file_groups.items():
            xlsx_path = group["xlsx_path"]
            output_path = output_dir / fname

            if not output_path.exists():
                prev_path = prev_job_dir / fname if prev_job_dir else None
                if prev_path and prev_path.exists():
                    shutil.copy2(prev_path, output_path)
                else:
                    shutil.copy2(xlsx_path, output_path)

            t0 = time.time()
            wb = load_workbook(output_path)
            print(f"[BibleTabling] 워크북 열기: {fname} ({len(wb.sheetnames)}시트, {time.time()-t0:.1f}s)")

            for edit in group["edits"]:
                t1 = time.time()
                sheet_name = edit.sheet or edit.table
                try:
                    result = self._execute_single_edit(edit, wb, xlsx_path)
                    is_csv = hasattr(edit, 'csv_set') and edit.csv_set
                    print(f"[BibleTabling]   {sheet_name}{' (CSV)' if is_csv else ''}: {result['rows_matched']}행 {result['cells_modified']}셀 ({time.time()-t1:.2f}s)")
                except Exception as e:
                    result = {
                        "table": edit.table, "file": xlsx_path.name,
                        "sheet": sheet_name, "error": str(e),
                        "rows_matched": 0, "cells_modified": 0, "changes": [],
                    }
                    print(f"[BibleTabling]   {sheet_name}: ❌ {e} ({time.time()-t1:.2f}s)")
                results.append(result)

            t2 = time.time()
            wb.save(output_path)
            wb.close()
            print(f"[BibleTabling] 저장 완료: {fname} ({len(group['edits'])}시트, save {time.time()-t2:.1f}s)")

        print(f"[BibleTabling] 전체 완료: {len(results)}개 편집, {time.time()-t_total:.1f}s")
        return results

    def apply_edits_streaming(self, edits: list, output_dir: Path, prev_job_dir: Path = None):
        """
        SSE 스트리밍용 제너레이터. 각 편집 단계마다 진행 이벤트를 yield.
        개별 편집 실패 시에도 계속 진행 (fault-tolerant).
        """
        file_groups = self._group_edits_by_file(edits)
        total_edits = sum(len(g["edits"]) for g in file_groups.values())
        results = []
        step = 0
        t_total = time.time()

        for fname, group in file_groups.items():
            xlsx_path = group["xlsx_path"]
            output_path = output_dir / fname

            if not output_path.exists():
                prev_path = prev_job_dir / fname if prev_job_dir else None
                if prev_path and prev_path.exists():
                    shutil.copy2(prev_path, output_path)
                else:
                    shutil.copy2(xlsx_path, output_path)

            t0 = time.time()
            wb = load_workbook(output_path)
            load_sec = time.time() - t0
            yield {
                "type": "file_open", "file": fname,
                "sheets": len(group["edits"]), "load_sec": round(load_sec, 1),
            }

            for edit in group["edits"]:
                step += 1
                sheet_name = edit.sheet or edit.table
                t1 = time.time()
                try:
                    result = self._execute_single_edit(edit, wb, xlsx_path)
                    result["status"] = "ok"
                    elapsed = round(time.time() - t1, 2)
                    print(f"[BibleTabling]   {sheet_name}: {result['rows_matched']}행 {result['cells_modified']}셀 ({elapsed}s)")
                    results.append(result)
                    yield {
                        "type": "edit_done", "step": step, "total": total_edits,
                        "result": result, "elapsed": elapsed,
                    }
                except Exception as e:
                    result = {
                        "table": edit.table, "file": xlsx_path.name,
                        "sheet": sheet_name, "status": "error", "error": str(e),
                        "rows_matched": 0, "cells_modified": 0, "changes": [],
                    }
                    print(f"[BibleTabling]   {sheet_name}: ❌ {e}")
                    results.append(result)
                    yield {
                        "type": "edit_error", "step": step, "total": total_edits,
                        "result": result,
                    }

            t2 = time.time()
            wb.save(output_path)
            wb.close()
            save_sec = round(time.time() - t2, 1)
            print(f"[BibleTabling] 저장 완료: {fname} (save {save_sec}s)")
            yield {"type": "file_saved", "file": fname, "save_sec": save_sec}

        total_sec = round(time.time() - t_total, 1)
        print(f"[BibleTabling] 전체 완료: {len(results)}개 편집, {total_sec}s")
        yield {"type": "complete", "results": results, "total_sec": total_sec}

    # ── 행 추가 ──────────────────────────────────────────────────────────────

    def clone_rows(self, table: str, file_hint: str, sheet_hint: str,
                   source_column: str, source_value: str,
                   override_csv: str | None,
                   output_dir: Path, prev_job_dir: Path = None) -> dict:
        """
        원본 행을 복제하여 새 행 추가. override_csv의 각 행마다 원본 전체를 복사하고
        지정된 컬럼만 오버라이드.
        override_csv가 없으면 원본을 1회 그대로 복제.
        override_csv에 N행이 있고 원본이 M행이면 N×M행 생성.
        """
        xlsx_path = self._find_xlsx_file(table, file_hint)
        sheet_name = sheet_hint or table

        output_path = output_dir / xlsx_path.name
        if not output_path.exists():
            prev_path = prev_job_dir / xlsx_path.name if prev_job_dir else None
            if prev_path and prev_path.exists():
                shutil.copy2(prev_path, output_path)
            else:
                shutil.copy2(xlsx_path, output_path)

        wb = load_workbook(output_path)
        ws = None
        for sn in wb.sheetnames:
            if sn.lower() == sheet_name.lower():
                ws = wb[sn]
                break
        if ws is None:
            wb.close()
            raise ValueError(f"시트 '{sheet_name}' 없음")

        header_row_idx = self._find_header_row(ws)
        header_map = self._get_header_map(ws, header_row_idx)
        # 역매핑: col_idx → header_name
        idx_to_name = {v: k for k, v in header_map.items() if k == k}  # 원본 케이스
        max_col = max(header_map.values()) if header_map else 0

        src_col_idx = header_map.get(source_column) or header_map.get(source_column.lower())
        if src_col_idx is None:
            wb.close()
            raise ValueError(f"소스 필터 컬럼 '{source_column}' 없음 (시트: {sheet_name})")

        source_rows: list[dict[int, any]] = []
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=src_col_idx).value
            if cell_val is not None and str(cell_val).strip().lower() == source_value.strip().lower():
                row_data = {}
                for col_idx in range(1, max_col + 1):
                    row_data[col_idx] = ws.cell(row=row_idx, column=col_idx).value
                source_rows.append(row_data)

        if not source_rows:
            wb.close()
            raise ValueError(
                f"소스 행 없음: {source_column}={source_value} (시트: {sheet_name})"
            )

        override_sets: list[dict[int, str]] = [{}]
        if override_csv:
            reader = csv.reader(io.StringIO(override_csv.strip()))
            ov_headers = [h.strip() for h in next(reader)]
            ov_col_indices = []
            for cn in ov_headers:
                idx = header_map.get(cn) or header_map.get(cn.lower())
                ov_col_indices.append(idx)

            override_sets = []
            for csv_row in reader:
                if not csv_row:
                    continue
                ov = {}
                for i, col_idx in enumerate(ov_col_indices):
                    if col_idx is None or i >= len(csv_row):
                        continue
                    ov[col_idx] = csv_row[i]
                override_sets.append(ov)

        if not override_sets:
            override_sets = [{}]

        rows_added = 0
        next_row = ws.max_row + 1
        sample_rows: list[dict[str, str]] = []

        for ov in override_sets:
            for src in source_rows:
                row_snapshot: dict[str, str] = {}
                for col_idx in range(1, max_col + 1):
                    if col_idx in ov:
                        val_str = ov[col_idx]
                        try:
                            v = float(val_str)
                            val = int(v) if v == int(v) else round(v, 4)
                        except (ValueError, TypeError):
                            val = val_str
                    else:
                        val = src.get(col_idx)
                    cell = ws.cell(row=next_row, column=col_idx, value=val)
                    cell.fill = AI_HIGHLIGHT
                    col_name = idx_to_name.get(col_idx, f"col{col_idx}")
                    if col_idx in ov:
                        row_snapshot[col_name] = str(val)
                if len(sample_rows) < 20:
                    sample_rows.append(row_snapshot)
                next_row += 1
                rows_added += 1

        wb.save(output_path)
        wb.close()

        # 오버라이드된 컬럼명 목록
        override_cols = []
        if override_csv:
            try:
                reader = csv.reader(io.StringIO(override_csv.strip()))
                override_cols = [h.strip() for h in next(reader)]
            except StopIteration:
                pass

        return {
            "table": table,
            "file": xlsx_path.name,
            "rows_added": rows_added,
            "source_rows": len(source_rows),
            "override_sets": len(override_sets),
            "override_columns": override_cols,
            "sample_rows": sample_rows,
        }

    def add_rows_csv(self, table: str, file_hint: str, sheet_hint: str,
                     csv_data: str, output_dir: Path, prev_job_dir: Path = None) -> dict:
        """CSV 문자열로 행 추가. 헤더 행 + 데이터 행 형식."""
        xlsx_path = self._find_xlsx_file(table, file_hint)
        sheet_name = sheet_hint or table

        output_path = output_dir / xlsx_path.name
        if not output_path.exists():
            prev_path = prev_job_dir / xlsx_path.name if prev_job_dir else None
            if prev_path and prev_path.exists():
                shutil.copy2(prev_path, output_path)
            else:
                shutil.copy2(xlsx_path, output_path)

        wb = load_workbook(output_path)
        ws = None
        for sn in wb.sheetnames:
            if sn.lower() == sheet_name.lower():
                ws = wb[sn]
                break
        if ws is None:
            wb.close()
            raise ValueError(f"시트 '{sheet_name}' 없음")

        header_row = self._find_header_row(ws)
        header_map = self._get_header_map(ws, header_row)

        reader = csv.reader(io.StringIO(csv_data.strip()))
        csv_headers = [h.strip() for h in next(reader)]

        col_indices = []
        for cn in csv_headers:
            idx = header_map.get(cn) or header_map.get(cn.lower())
            col_indices.append(idx)

        rows_added = 0
        next_row = ws.max_row + 1

        for csv_row in reader:
            if not csv_row:
                continue
            for i, col_idx in enumerate(col_indices):
                if col_idx is None or i >= len(csv_row):
                    continue
                val_str = csv_row[i]
                try:
                    v = float(val_str)
                    val = int(v) if v == int(v) else round(v, 4)
                except (ValueError, TypeError):
                    val = val_str
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                cell.fill = AI_HIGHLIGHT
            next_row += 1
            rows_added += 1

        wb.save(output_path)
        wb.close()

        return {
            "table": table,
            "file": xlsx_path.name,
            "rows_added": rows_added,
        }

    def add_rows(self, table: str, file_hint: str, sheet_hint: str,
                 rows: list[dict], output_dir: Path, prev_job_dir: Path = None) -> dict:
        """테이블에 새 행 추가 (노란색 하이라이트)"""
        xlsx_path = self._find_xlsx_file(table, file_hint)
        sheet_name = sheet_hint or table

        output_path = output_dir / xlsx_path.name
        if not output_path.exists():
            prev_path = prev_job_dir / xlsx_path.name if prev_job_dir else None
            if prev_path and prev_path.exists():
                shutil.copy2(prev_path, output_path)
            else:
                shutil.copy2(xlsx_path, output_path)

        wb = load_workbook(output_path)
        ws = None
        for sn in wb.sheetnames:
            if sn.lower() == sheet_name.lower():
                ws = wb[sn]
                break

        if ws is None:
            wb.close()
            raise ValueError(f"시트 '{sheet_name}'을 찾을 수 없습니다.")

        header_row = self._find_header_row(ws)
        header_map = self._get_header_map(ws, header_row)

        rows_added = 0
        next_row = ws.max_row + 1

        for row_data in rows:
            for col_name, value in row_data.items():
                col_idx = header_map.get(col_name) or header_map.get(col_name.lower())
                if col_idx is None:
                    continue
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.fill = AI_HIGHLIGHT
            next_row += 1
            rows_added += 1

        wb.save(output_path)
        wb.close()

        return {
            "table": table,
            "file": xlsx_path.name,
            "rows_added": rows_added,
        }

    # ── Preview: AI 하이라이트 행 추출 ────────────────────────────────────────

    @staticmethod
    def preview_job(job_dir: Path) -> list[dict]:
        """job_dir의 모든 xlsx에서 AI_HIGHLIGHT(노란색) 셀이 있는 행만 추출."""
        result = []
        for xlsx_path in sorted(job_dir.glob("*.xlsx")):
            if xlsx_path.name.startswith("~$"):
                continue
            try:
                wb = load_workbook(xlsx_path, data_only=False)
            except Exception:
                continue
            file_entry: dict = {"filename": xlsx_path.name, "sheets": []}
            for ws in wb.worksheets:
                if ws.title.lower() in META_SHEET_NAMES or '#' in ws.title:
                    continue
                best_idx, best_count = 1, 0
                for ri in range(1, min(5, ws.max_row + 1)):
                    cnt = sum(1 for c in ws[ri] if c.value is not None and str(c.value).strip())
                    if cnt > best_count:
                        best_count = cnt
                        best_idx = ri
                header_row_idx = best_idx
                headers: list[str] = []
                col_to_name: dict[int, str] = {}
                for cell in ws[header_row_idx]:
                    if cell.value is not None:
                        name = str(cell.value).strip()
                        if name:
                            headers.append(name)
                            col_to_name[cell.column] = name
                if not headers:
                    continue
                max_col = max(col_to_name.keys()) if col_to_name else 0

                highlighted_rows: list[dict] = []
                for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                    has_highlight = False
                    row_cells: dict[str, dict] = {}
                    for col_idx in range(1, max_col + 1):
                        col_name = col_to_name.get(col_idx)
                        if not col_name:
                            continue
                        cell = ws.cell(row=row_idx, column=col_idx)
                        is_hl = (
                            cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb
                            and str(cell.fill.fgColor.rgb).upper().endswith('FFFF00')
                        )
                        if is_hl:
                            has_highlight = True
                        val = cell.value
                        if val is None:
                            val = ""
                        elif isinstance(val, float) and val == int(val):
                            val = int(val)
                        row_cells[col_name] = {"value": val, "highlighted": bool(is_hl)}
                    if has_highlight:
                        highlighted_rows.append({"rowIndex": row_idx, "cells": row_cells})
                if highlighted_rows:
                    file_entry["sheets"].append({
                        "name": ws.title,
                        "headers": headers,
                        "rows": highlighted_rows,
                    })
            wb.close()
            if file_entry["sheets"]:
                result.append(file_entry)
        return result

    # ── Update Cells: 사용자 인라인 편집 저장 ─────────────────────────────────

    @staticmethod
    def update_cells(job_dir: Path, filename: str, sheet_name: str,
                     updates: list[dict]) -> dict:
        """job_dir 내 xlsx의 특정 셀 값을 업데이트."""
        xlsx_path = job_dir / filename
        if not xlsx_path.exists():
            raise FileNotFoundError(f"파일 없음: {filename}")

        wb = load_workbook(xlsx_path)
        ws = None
        for sn in wb.sheetnames:
            if sn.lower() == sheet_name.lower():
                ws = wb[sn]
                break
        if ws is None:
            wb.close()
            raise ValueError(f"시트 없음: {sheet_name}")

        best_idx, best_count = 1, 0
        for ri in range(1, min(5, ws.max_row + 1)):
            cnt = sum(1 for c in ws[ri] if c.value is not None and str(c.value).strip())
            if cnt > best_count:
                best_count = cnt
                best_idx = ri
        header_map: dict[str, int] = {}
        for cell in ws[best_idx]:
            if cell.value is not None:
                name = str(cell.value).strip()
                if name:
                    header_map[name.lower()] = cell.column
                    header_map[name] = cell.column

        user_fill = PatternFill(start_color='FF90CAF9', end_color='FF90CAF9', fill_type='solid')
        updated = 0
        for upd in updates:
            row = upd.get("row")
            col_name = upd.get("column", "")
            new_val = upd.get("value", "")
            col_idx = header_map.get(col_name) or header_map.get(col_name.lower())
            if row is None or col_idx is None:
                continue
            cell = ws.cell(row=row, column=col_idx)
            try:
                v = float(new_val)
                cell.value = int(v) if v == int(v) else round(v, 4)
            except (ValueError, TypeError):
                cell.value = new_val
            cell.fill = user_fill
            updated += 1

        wb.save(xlsx_path)
        wb.close()
        return {"updated": updated, "file": filename, "sheet": sheet_name}
